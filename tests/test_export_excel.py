"""The workbook: what the screen shows, in a form that can be calculated with."""

import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from agents.sme_questions import SmeQuestion, SmeQuestionProposal
from analysis.report import build_report
from analysis.summary import build_summary
from core import palette
from core.table import load_table
from export.excel import MONEY_FORMAT, build_workbook
from tests.conftest import FakeClient

BRAND = palette.BRAND.lstrip("#").upper()


@pytest.fixture
def workbook(lever_run):
    build_summary(
        lever_run,
        client=FakeClient(
            SmeQuestionProposal(
                questions=[
                    SmeQuestion(
                        question="Is the missing purchase order a policy choice?",
                        rationale="Most bookings carry none.",
                        addressee="procurement",
                        unlocks="Whether maverick spend is a gap or the norm.",
                    )
                ]
            )
        ),
    )
    data = build_workbook(build_report(lever_run), load_table(lever_run))
    return load_workbook(io.BytesIO(data)), lever_run


def test_the_workbook_opens_on_a_cover_and_ends_on_the_data(workbook):
    book, _ = workbook

    assert book.sheetnames[0] == "Cover"
    assert book.sheetnames[-2:] == ["Data", "Data (full)"]
    assert "Overview" in book.sheetnames


def test_every_figure_gets_its_own_sheet_with_its_chart(workbook):
    book, _ = workbook

    charted = [name for name in book.sheetnames if book[name]._charts]
    assert charted, "no sheet carries a chart"
    for name in charted:
        sheet = book[name]
        # A chart without its numbers beside it cannot be checked by a reader.
        assert sheet.max_row > len(sheet._charts)


def test_amounts_are_numbers_a_pivot_can_use(workbook):
    book, run_id = workbook
    sheet = book["Data (full)"]

    header = {cell.value: cell.column for cell in sheet[1]}
    cell = sheet.cell(row=2, column=header["amount_eur"])

    assert isinstance(cell.value, (int, float))
    assert cell.number_format == MONEY_FORMAT


def test_the_full_data_sheet_holds_every_row_and_column(workbook):
    book, run_id = workbook
    table = load_table(run_id)
    sheet = book["Data (full)"]

    assert sheet.max_row == len(table) + 1
    assert sheet.max_column == len(table.columns)
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref


def test_the_business_data_sheet_leaves_out_the_flag_columns(workbook):
    book, _ = workbook
    header = [cell.value for cell in book["Data"][1]]

    assert not [name for name in header if str(name).startswith("flag_")]
    # But keeps the eligibility columns, or nobody can see why a row is excluded.
    assert [name for name in header if str(name).startswith("include_")]


def test_the_waterfall_is_built_on_an_invisible_base_series(workbook):
    book, run_id = workbook
    sheet = book["From booked to negotiable"]
    chart = sheet._charts[0]

    assert chart.grouping == "stacked"
    base, step = chart.series
    assert base.graphicalProperties.noFill is True
    # One coloured point per step of the chain, not one for the whole series.
    from analysis.spend_report import spend_chain

    assert len(step.data_points) == len(spend_chain(load_table(run_id)).chain)


def test_chart_colours_come_from_the_validated_palette(workbook):
    book, _ = workbook
    allowed = {
        colour.lstrip("#").upper()
        for colour in (
            *palette.CATEGORICAL,
            palette.BRAND,
            palette.WATERFALL_DEDUCTION,
            palette.WATERFALL_RESULT,
            palette.OTHER,
        )
    }

    used = set()
    for name in book.sheetnames:
        for chart in book[name]._charts:
            for series in chart.series:
                fill = series.graphicalProperties.solidFill
                if fill is not None and getattr(fill, "srgbClr", None):
                    used.add(fill.srgbClr)
                for point in series.data_points:
                    used.add(point.graphicalProperties.solidFill.srgbClr)

    assert used, "no chart colours found to check"
    assert used <= allowed, f"colours outside the palette: {used - allowed}"


def test_the_brand_colour_carries_the_headers(workbook):
    book, _ = workbook
    sheet = book["Data (full)"]

    assert BRAND in str(sheet["A1"].fill.fgColor.rgb)
    assert sheet["A1"].font.bold


def test_a_run_with_no_table_still_produces_a_readable_workbook(defective_run):
    from analysis.report import build_report as build

    data = build_workbook(build(defective_run), pd.DataFrame())
    book = load_workbook(io.BytesIO(data))

    assert book.sheetnames == ["Cover"]


def test_the_export_is_kept_with_the_run_it_came_from(lever_run):
    """A report handed to someone stays findable next to its evidence."""
    from core.run import load_run, step_path
    from export.artifacts import PAGE_NAME, STEP, WORKBOOK_NAME, build_exports, has_exports

    assert not has_exports(lever_run)

    workbook, page = build_exports(lever_run)

    target = step_path(lever_run, STEP)
    assert (target / WORKBOOK_NAME).read_bytes() == workbook
    assert (target / PAGE_NAME).read_text(encoding="utf-8") == page
    assert STEP in [step.step for step in load_run(lever_run).steps]


def test_the_file_name_says_who_and_when(lever_run):
    from analysis.report import build_report as build
    from export.artifacts import file_stem

    stem = file_stem(build(lever_run))

    assert stem.startswith("procurement_levers_")
    assert stem[-8:].isdigit()  # the date, so two exports never collide silently
