"""Deterministic shape of every sheet in an upload.

An ERP data request usually comes back as a workbook, not a table: a cover
letter, filling instructions, the actual transactions, and small lookup tables
for FX rates or suppliers. Which sheet is which is mostly a question of shape,
and shape can be measured without asking a model anything.

Measured on a real submission:

    sheet                  header  fill  rect   table
    1. Brief                  no   0.89  0.85     no
    2. How to Submit          no   0.63  0.28     no
    3. Spend Data            yes   0.91  1.00    yes
    4. Supplier Master       yes   0.89  1.00    yes
    5. FX                    yes   1.00  1.00    yes

The header check carries the decision on the prose sheets, and it is a real
signal rather than a lucky one: their first row is a title filling one of two
columns, whereas a data table's header fills every column it spans. Fill ratio
and rectangularity would both have kept the cover letter on their own -- it
reaches 0.85 rectangularity simply by being narrow.

Telling an FX table apart from a supplier master is a question of meaning, not
shape -- that is what the triage agent is for.
"""

import datetime
import re

from openpyxl import load_workbook

from core.config import MAX_SAMPLE_LENGTH
from core.models import FileFormat, ReadOptions, SheetProfile
from ingestion.readers import read_with_options

# Enough rows to judge the shape of a sheet without scanning millions of cells.
PROFILE_ROWS = 40
SAMPLE_ROWS = 3

MIN_FILL_RATIO = 0.5
MIN_RECTANGULARITY = 0.8
# A single column, or a header with nothing under it, is not a table worth analysing.
MIN_COLUMNS = 2
# Share of the header width a row must fill to count as part of the table body.
ROW_WIDTH_TOLERANCE = 0.6

_DATE_TEXT = re.compile(r"^\d{4}-\d{2}-\d{2}|^\d{1,2}[./-]\d{1,2}[./-]\d{4}")


def profile_sheets(
    data: bytes, fmt: FileFormat, read_options: ReadOptions
) -> list[SheetProfile]:
    if fmt == "csv":
        frame = read_with_options(data, fmt, read_options)
        rows = [tuple(frame.columns)] + [
            tuple(row) for row in frame.head(PROFILE_ROWS).itertuples(index=False)
        ]
        return [_profile("", rows, total_rows=len(frame) + 1)]

    workbook = load_workbook(data_stream(data), read_only=True)
    try:
        return [
            _profile(
                name,
                list(workbook[name].iter_rows(max_row=PROFILE_ROWS, values_only=True)),
                total_rows=workbook[name].max_row or 0,
            )
            for name in workbook.sheetnames
        ]
    finally:
        workbook.close()


def data_stream(data: bytes):
    import io

    return io.BytesIO(data)


def best_table_sheet(profiles: list[SheetProfile]) -> str | None:
    """The sheet most likely to hold the transactions, decided without a model.

    Used to preview something useful before the agent has run. Falls back to the
    largest table, then to nothing at all.
    """
    tables = [profile for profile in profiles if profile.looks_like_table]
    if not tables:
        return None
    transactional = [p for p in tables if p.has_date_column and p.has_numeric_column]
    return max(transactional or tables, key=lambda profile: profile.rows).name or None


def _profile(name: str, raw_rows: list[tuple], total_rows: int) -> SheetProfile:
    rows = [row for row in raw_rows if _filled(row)]
    if not rows:
        return SheetProfile(
            name=name,
            rows=0,
            columns=0,
            fill_ratio=0.0,
            rectangularity=0.0,
            has_header_row=False,
            has_numeric_column=False,
            has_date_column=False,
            looks_like_table=False,
            header=[],
            sample_rows=[],
        )

    width = max(_filled(row) for row in rows)
    header = [_text(cell) for cell in rows[0] if _text(cell)]
    body = rows[1:]

    has_header = (
        len(header) == width
        and len(set(header)) == len(header)
        and all(isinstance(cell, str) for cell in rows[0] if cell is not None)
    )
    rectangularity = (
        sum(1 for row in body if _filled(row) >= width * ROW_WIDTH_TOLERANCE) / len(body)
        if body
        else 0.0
    )
    fill_ratio = sum(_filled(row) for row in rows) / (len(rows) * width)

    return SheetProfile(
        name=name,
        rows=total_rows,
        columns=width,
        fill_ratio=round(fill_ratio, 3),
        rectangularity=round(rectangularity, 3),
        has_header_row=has_header,
        has_numeric_column=any(_is_number(cell) for row in body for cell in row),
        has_date_column=any(_is_date(cell) for row in body for cell in row),
        looks_like_table=(
            has_header
            and bool(body)
            and width >= MIN_COLUMNS
            and fill_ratio >= MIN_FILL_RATIO
            and rectangularity >= MIN_RECTANGULARITY
        ),
        header=header,
        sample_rows=[
            [_text(cell)[:MAX_SAMPLE_LENGTH] for cell in row] for row in body[:SAMPLE_ROWS]
        ],
    )


def _filled(row: tuple) -> int:
    return sum(1 for cell in row if _text(cell))


def _text(cell) -> str:
    return "" if cell is None else str(cell).strip()


def _is_number(cell) -> bool:
    return isinstance(cell, (int, float)) and not isinstance(cell, bool)


def _is_date(cell) -> bool:
    if isinstance(cell, (datetime.date, datetime.datetime)):
        return True
    return isinstance(cell, str) and bool(_DATE_TEXT.match(cell.strip()))
