"""Parse ERP exports into raw text DataFrames.

Every value is read as text. Type inference belongs to the deterministic rule
engine further down the pipeline -- doing it here would silently strip leading
zeros from supplier IDs and misread German decimal formats such as ``1.250,00``.
"""

import csv
import io
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from core.config import ALLOWED_EXTENSIONS
from core.models import FileFormat, ReadOptions

ENCODING_CANDIDATES = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
DELIMITER_CANDIDATES = ",;\t|"
SNIFF_BYTES = 64 * 1024


class UnsupportedFileError(ValueError):
    """Raised for file types the ingestion layer does not handle."""


def file_format(filename: str) -> FileFormat:
    suffix = Path(filename).suffix.lower().lstrip(".")
    if suffix not in ALLOWED_EXTENSIONS:
        raise UnsupportedFileError(
            f"{filename!r} has unsupported extension {suffix!r}; "
            f"expected one of {', '.join(ALLOWED_EXTENSIONS)}"
        )
    return suffix  # type: ignore[return-value]


def list_sheets(data: bytes) -> list[str]:
    workbook = load_workbook(io.BytesIO(data), read_only=True)
    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def read_tabular(
    data: bytes, filename: str, sheet: str | None = None
) -> tuple[pd.DataFrame, ReadOptions]:
    """Parse an upload, detecting whatever the format does not state explicitly."""
    fmt = file_format(filename)
    options = detect_options(data, fmt, sheet)
    return read_with_options(data, fmt, options), options


def read_with_options(data: bytes, fmt: FileFormat, options: ReadOptions) -> pd.DataFrame:
    """Parse using known options, so a stored file always yields the same frame."""
    if fmt == "csv":
        text = data.decode(options.encoding or "utf-8")
        return pd.read_csv(
            io.StringIO(text),
            sep=options.delimiter,
            dtype=str,
            keep_default_na=False,
        )
    return pd.read_excel(
        io.BytesIO(data),
        sheet_name=options.sheet,
        dtype=str,
        keep_default_na=False,
        engine="openpyxl",
    )


def detect_options(data: bytes, fmt: FileFormat, sheet: str | None = None) -> ReadOptions:
    """Work out how to parse this file. For CSV that is encoding and delimiter."""
    if fmt == "csv":
        encoding = _detect_encoding(data)
        return ReadOptions(encoding=encoding, delimiter=_detect_delimiter(data.decode(encoding)))

    sheets = list_sheets(data)
    if sheet is not None and sheet not in sheets:
        raise ValueError(f"sheet {sheet!r} not found; available: {', '.join(sheets)}")
    return ReadOptions(sheet=sheet or sheets[0])


def file_options(data: bytes, fmt: FileFormat) -> ReadOptions:
    """Options that belong to the file rather than to one sheet of it.

    Which sheet to read is decided later by triage, so it is deliberately left
    unset here.
    """
    return ReadOptions() if fmt == "xlsx" else detect_options(data, fmt)


def _detect_encoding(data: bytes) -> str:
    for encoding in ENCODING_CANDIDATES:
        try:
            data.decode(encoding)
        except UnicodeDecodeError:
            continue
        return encoding
    raise UnsupportedFileError("file could not be decoded with any supported encoding")


def _detect_delimiter(text: str) -> str:
    try:
        return csv.Sniffer().sniff(text[:SNIFF_BYTES], delimiters=DELIMITER_CANDIDATES).delimiter
    except csv.Error:
        # Single-column exports give the sniffer nothing to separate.
        return ","
