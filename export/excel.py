"""The analysis as a workbook, for the reader who carries on calculating.

Charts are native Excel charts bound to the sheet they sit on, so they can be
recoloured, resized and pasted into a deck. Amounts are written as numbers with a
display format, never as pre-formatted text: a workbook whose figures cannot be
pivoted misses the point of being a workbook.

Everything drawn here comes from analysis.report, so the workbook shows what the
screen shows.
"""

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, Series
from openpyxl.chart.marker import DataPoint
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from analysis.report import Block, ReportDocument
from core import palette

# Column labels ending in these hold money; shares are fractions of one. The
# same convention the screens use, so one rule covers every table.
AMOUNT_SUFFIXES = ("(EUR)", "(local)")
SHARE_COLUMNS = ("share", "cumulative", "Share of net", "PO coverage", "Without contract")

MONEY_FORMAT = "#,##0"
SHARE_FORMAT = "0.0%"

VISUALS = "Visuals"

CHART_WIDTH = 22
CHART_HEIGHT = 10
MAX_COLUMN_WIDTH = 60


def _rgb(colour: str) -> str:
    """openpyxl wants RRGGBB without the hash."""
    return colour.lstrip("#").upper()


HEADER_FILL = PatternFill("solid", fgColor=_rgb(palette.BRAND))
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BAND_FILL = PatternFill("solid", fgColor=_rgb(palette.SEQUENTIAL[0]))
TITLE_FONT = Font(color=_rgb(palette.BRAND), bold=True, size=16)
SECTION_FONT = Font(color=_rgb(palette.BRAND), bold=True, size=12)
MUTED_FONT = Font(color=_rgb(palette.TEXT_SECONDARY), size=10)
THIN_BOTTOM = Border(bottom=Side(style="thin", color=_rgb(palette.GRID)))


def build_workbook(document: ReportDocument, table: pd.DataFrame) -> bytes:
    """The whole report as xlsx bytes, ready to hand to a download button."""
    workbook = Workbook()
    workbook.remove(workbook.active)

    _cover_sheet(workbook, document)
    for section in document.sections:
        if section.title == VISUALS:
            # One chart per tab beats fourteen stacked on one. The assembly is
            # unchanged; only how it is laid out differs, which is the renderer's
            # job.
            for block in section.blocks:
                _block_sheet(workbook, block)
        else:
            _section_sheet(workbook, section)
    _data_sheets(workbook, table)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --- the sheets ------------------------------------------------------------


def _cover_sheet(workbook: Workbook, document: ReportDocument) -> None:
    cover = document.cover
    sheet = workbook.create_sheet("Cover")
    sheet.sheet_view.showGridLines = False

    _write(sheet, 2, 2, cover.title, TITLE_FONT)
    _write(sheet, 3, 2, cover.group, SECTION_FONT)

    row = 5
    for label, value in cover.metrics:
        _write(sheet, row, 2, label, MUTED_FONT)
        _write(sheet, row, 3, value, Font(bold=True, size=12))
        row += 1

    row += 1
    facts = [
        ("Prepared on", cover.prepared_on.isoformat()),
        ("Run", cover.run_id),
        ("Source files", ", ".join(cover.sources) or "-"),
        ("Rows", f"{cover.rows_total:,}"),
        ("Rows entering the analysis", f"{cover.rows_analysed:,}"),
    ]
    for label, value in facts:
        _write(sheet, row, 2, label, MUTED_FONT)
        _write(sheet, row, 3, value)
        row += 1

    row += 1
    _write(sheet, row, 2, cover.note, MUTED_FONT)
    sheet.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.merge_cells(start_row=row, start_column=2, end_row=row + 2, end_column=7)

    sheet.column_dimensions["A"].width = 2
    sheet.column_dimensions["B"].width = 30
    sheet.column_dimensions["C"].width = 46


def _section_sheet(workbook: Workbook, section) -> None:
    sheet = _new_sheet(workbook, section.title)
    _write(sheet, 1, 1, section.title, TITLE_FONT)

    row = 3
    for block in section.blocks:
        row = _write_block(sheet, row, block)
    _fit_columns(sheet)


def _block_sheet(workbook: Workbook, block: Block) -> None:
    sheet = _new_sheet(workbook, block.title)
    _write_block(sheet, 1, block)
    _fit_columns(sheet)


def _new_sheet(workbook: Workbook, title: str) -> Worksheet:
    sheet = workbook.create_sheet(_sheet_name(title, workbook.sheetnames))
    sheet.sheet_view.showGridLines = False
    return sheet


def _write_block(sheet: Worksheet, row: int, block: Block) -> int:
    _write(sheet, row, 1, block.title, SECTION_FONT)
    row += 1
    if block.caption:
        _write(sheet, row, 1, block.caption, MUTED_FONT)
        row += 1

    if block.metrics:
        for offset, (label, value) in enumerate(block.metrics):
            _write(sheet, row, 1 + offset * 2, label, MUTED_FONT)
            _write(sheet, row + 1, 1 + offset * 2, value, Font(bold=True))
        row += 3

    if block.body:
        for line in block.body.splitlines():
            if line.strip():
                _write(sheet, row, 1, line.lstrip("- "))
                row += 1
        row += 1

    if block.table is not None and not block.table.empty:
        row = _write_table(sheet, row, block.table) + 1

    for figure in block.figures:
        row = _write_figure(sheet, row, figure) + 1

    return row + 1


def _write_table(sheet: Worksheet, row: int, frame: pd.DataFrame, banded: bool = True) -> int:
    """A styled table starting at `row`. Returns the last row written."""
    for offset, column in enumerate(frame.columns, start=1):
        cell = sheet.cell(row=row, column=offset, value=str(column))
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT

    formats = {
        offset: _number_format(column)
        for offset, column in enumerate(frame.columns, start=1)
    }
    for index, record in enumerate(frame.itertuples(index=False, name=None)):
        target = row + 1 + index
        for offset, value in enumerate(record, start=1):
            cell = sheet.cell(row=target, column=offset, value=_cell_value(value))
            cell.border = THIN_BOTTOM
            if banded and index % 2:
                cell.fill = BAND_FILL
            if formats[offset]:
                cell.number_format = formats[offset]
    return row + len(frame)


def _write_figure(sheet: Worksheet, row: int, figure) -> int:
    """The chart's own numbers, then the chart drawn from them."""
    frame = figure.data
    if frame.empty:
        return row
    if figure.caption:
        _write(sheet, row, 1, figure.caption, MUTED_FONT)
        row += 1

    header = row
    last = _write_table(sheet, header, frame, banded=False)
    chart = _chart_for(frame, sheet, header, last)
    if chart is not None:
        chart.width, chart.height = CHART_WIDTH, CHART_HEIGHT
        sheet.add_chart(chart, f"{get_column_letter(len(frame.columns) + 2)}{header}")
        return max(last, header + 20)
    return last


def _data_sheets(workbook: Workbook, table: pd.DataFrame) -> None:
    if table.empty:
        return
    from core.canonical import CANONICAL_FIELDS

    business = [f.key for f in CANONICAL_FIELDS if f.key in table.columns]
    business += [
        column
        for column in ("amount_eur", "supplier_normalized", "company_normalized")
        if column in table.columns
    ]
    business += [c for c in table.columns if c.startswith("include_")]
    _raw_sheet(workbook, "Data", table[business])
    _raw_sheet(workbook, "Data (full)", table)


def _raw_sheet(workbook: Workbook, title: str, frame: pd.DataFrame) -> None:
    """Rows in bulk. write_only keeps 18k x 74 under six seconds."""
    sheet = workbook.create_sheet(title)
    sheet.append([str(column) for column in frame.columns])
    for cell in sheet[1]:
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT

    numeric = {
        offset
        for offset, column in enumerate(frame.columns, start=1)
        if pd.api.types.is_numeric_dtype(frame[column])
    }
    for record in frame.itertuples(index=False, name=None):
        sheet.append([_cell_value(value) for value in record])
    for offset in numeric:
        letter = get_column_letter(offset)
        for cell in sheet[letter][1:]:
            cell.number_format = MONEY_FORMAT

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = (
        f"A1:{get_column_letter(len(frame.columns))}{len(frame) + 1}"
    )
    _fit_columns(sheet, sample=200)


# --- the charts ------------------------------------------------------------


def _chart_for(frame: pd.DataFrame, sheet: Worksheet, header: int, last: int):
    """Pick the native chart the figure's own columns call for."""
    columns = list(frame.columns)
    if {"start", "end", "step"} <= set(columns):
        return _waterfall(frame, sheet, header, last)
    if "month" in columns:
        return _line(frame, sheet, header, last, "spend", "month")
    if "status" in columns and "company" in columns:
        return _stacked_by_status(frame, sheet, header, last)
    if "supplier" in columns and "share" in columns and len(frame) <= palette.PIE_SLICES + 1:
        return _pie(frame, sheet, header, last, "spend", "supplier")
    for value, label in (("spend", "supplier"), ("spend", "company"), ("spend", "lever")):
        if value in columns and label in columns:
            return _bar(frame, sheet, header, last, value, label)
    return None


def _column(frame: pd.DataFrame, name: str) -> int:
    return list(frame.columns).index(name) + 1


def _waterfall(frame: pd.DataFrame, sheet: Worksheet, header: int, last: int):
    """Excel has no waterfall openpyxl can write, so it is a stacked bar.

    The lower series is the invisible run-up to each step and the upper series is
    the visible piece. Standard construction, readable in any Excel since 2010.
    """
    base_col = len(frame.columns) + 1
    size_col = base_col + 1
    sheet.cell(row=header, column=base_col, value="base").font = HEADER_FONT
    sheet.cell(row=header, column=base_col).fill = HEADER_FILL
    sheet.cell(row=header, column=size_col, value="size").font = HEADER_FONT
    sheet.cell(row=header, column=size_col).fill = HEADER_FILL

    for index, record in enumerate(frame.itertuples(index=False), start=1):
        low, high = sorted((float(record.start), float(record.end)))
        sheet.cell(row=header + index, column=base_col, value=low).number_format = MONEY_FORMAT
        sheet.cell(row=header + index, column=size_col, value=high - low).number_format = MONEY_FORMAT

    chart = BarChart()
    chart.type, chart.grouping, chart.overlap = "col", "stacked", 100
    chart.title = "From booked to negotiable"
    chart.y_axis.title = "EUR"

    invisible = Series(
        Reference(sheet, min_col=base_col, min_row=header, max_row=last), title="base"
    )
    invisible.graphicalProperties.noFill = True
    invisible.graphicalProperties.line.noFill = True
    chart.series.append(invisible)

    visible = Series(
        Reference(sheet, min_col=size_col, min_row=header, max_row=last), title="step"
    )
    roles = list(frame["role"]) if "role" in frame.columns else []
    for index, role in enumerate(roles):
        point = DataPoint(idx=index)
        point.graphicalProperties.solidFill = _rgb(
            {
                "deduction": palette.WATERFALL_DEDUCTION,
                "result": palette.WATERFALL_RESULT,
            }.get(role, palette.WATERFALL_TOTAL)
        )
        visible.data_points.append(point)
    chart.series.append(visible)

    chart.set_categories(
        Reference(sheet, min_col=_column(frame, "step"), min_row=header + 1, max_row=last)
    )
    chart.legend = None
    return chart


def _bar(frame, sheet, header, last, value: str, label: str):
    chart = BarChart()
    chart.type = "bar"
    chart.add_data(
        Reference(sheet, min_col=_column(frame, value), min_row=header, max_row=last),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=_column(frame, label), min_row=header + 1, max_row=last)
    )
    chart.series[0].graphicalProperties.solidFill = _rgb(palette.CATEGORICAL[0])
    chart.legend = None
    return chart


def _line(frame, sheet, header, last, value: str, label: str):
    chart = LineChart()
    chart.add_data(
        Reference(sheet, min_col=_column(frame, value), min_row=header, max_row=last),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=_column(frame, label), min_row=header + 1, max_row=last)
    )
    chart.series[0].graphicalProperties.line.solidFill = _rgb(palette.CATEGORICAL[0])
    chart.series[0].graphicalProperties.line.width = 20000
    chart.legend = None
    return chart


def _pie(frame, sheet, header, last, value: str, label: str):
    chart = PieChart()
    chart.add_data(
        Reference(sheet, min_col=_column(frame, value), min_row=header, max_row=last),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=_column(frame, label), min_row=header + 1, max_row=last)
    )
    slices = palette.categorical(min(len(frame), len(palette.CATEGORICAL)))
    for index in range(len(frame)):
        point = DataPoint(idx=index)
        colour = slices[index] if index < len(slices) else palette.OTHER
        point.graphicalProperties.solidFill = _rgb(colour)
        chart.series[0].data_points.append(point)
    return chart


def _stacked_by_status(frame: pd.DataFrame, sheet: Worksheet, header: int, last: int):
    """Contract coverage per company: one series per status, side by side.

    The long frame is pivoted into a block beside the table, because an Excel
    chart reads columns and the source here is one row per company and status.
    """
    wide = frame.pivot_table(
        index="company", columns="status", values="spend", aggfunc="sum", fill_value=0
    )
    start = len(frame.columns) + 2
    sheet.cell(row=header, column=start, value="company").font = HEADER_FONT
    for offset, status in enumerate(wide.columns, start=1):
        cell = sheet.cell(row=header, column=start + offset, value=str(status))
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
    for index, (company, record) in enumerate(wide.iterrows(), start=1):
        sheet.cell(row=header + index, column=start, value=str(company))
        for offset, value in enumerate(record, start=1):
            cell = sheet.cell(row=header + index, column=start + offset, value=float(value))
            cell.number_format = MONEY_FORMAT

    bottom = header + len(wide)
    chart = BarChart()
    chart.type, chart.grouping, chart.overlap = "bar", "stacked", 100
    chart.add_data(
        Reference(sheet, min_col=start + 1, max_col=start + len(wide.columns),
                  min_row=header, max_row=bottom),
        titles_from_data=True,
    )
    chart.set_categories(
        Reference(sheet, min_col=start, min_row=header + 1, max_row=bottom)
    )
    for index, series in enumerate(chart.series):
        series.graphicalProperties.solidFill = _rgb(palette.CATEGORICAL[index % 6])
    return chart


# --- small helpers ---------------------------------------------------------


def _write(sheet: Worksheet, row: int, column: int, value, font: Font | None = None):
    cell = sheet.cell(row=row, column=column, value=value)
    if font is not None:
        cell.font = font
    return cell


def _cell_value(value):
    """Numbers stay numbers. Everything Excel cannot hold becomes text."""
    if value is None or isinstance(value, (int, float, str, bool)):
        return value
    if pd.isna(value):
        return None
    if hasattr(value, "item"):
        return value.item()
    return str(value)


def _number_format(column) -> str | None:
    label = str(column)
    if label in SHARE_COLUMNS:
        return SHARE_FORMAT
    if label.endswith(AMOUNT_SUFFIXES) or label in ("spend", "amount", "start", "end", "EUR"):
        return MONEY_FORMAT
    return None


def _fit_columns(sheet: Worksheet, sample: int = 60) -> None:
    widths: dict[int, int] = {}
    for row in sheet.iter_rows(max_row=sample):
        for cell in row:
            if cell.value is not None:
                widths[cell.column] = max(widths.get(cell.column, 10), len(str(cell.value)) + 2)
    for column, width in widths.items():
        sheet.column_dimensions[get_column_letter(column)].width = min(width, MAX_COLUMN_WIDTH)


def _sheet_name(title: str, taken: list[str] | None = None) -> str:
    """Excel allows 31 characters, forbids []:*?/\\, and wants them unique."""
    cleaned = "".join(character for character in title if character not in "[]:*?/\\")
    name = cleaned[:31] or "Sheet"
    taken = taken or []
    if name not in taken:
        return name
    for suffix in range(2, 100):
        candidate = f"{name[:31 - len(str(suffix)) - 1]} {suffix}"
        if candidate not in taken:
            return candidate
    return name
